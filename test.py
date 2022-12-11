from openpyxl import Workbook, load_workbook
import os

# فتح القاعدة
database = load_workbook(filename="القاعدة\قاعدة الفصل الأول - ١٤٤٤هـ.xlsx")


# ورقة الحفظ بنات


def hefeth_G(database, from_row, to_row):
    hefeth_sheet = database['حفظ']

    for value in hefeth_sheet.iter_rows(min_row=from_row,
                                        max_row=to_row,
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        # تجواز الأميات
        if school == "أمية":
            continue

        test_template = load_workbook(
            filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                if str(track) == "1-ب" and i == 1:
                    continue
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['G4'] = track
            sheet['G5'] = level
            sheet['I4'] = id

        # file path creation
        isExist = os.path.exists(f"حفظ\{dar}")

        if not isExist:
            os.makedirs(f"حفظ\{dar}")

        # save
        test_template.save(filename=f"حفظ\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - حفظ")


# ورقة التعاهد بنات


def taahod_G(database, from_row, to_row):
    taahod_sheet = database['تعاهد']

    for value in taahod_sheet.iter_rows(min_row=from_row,
                                        max_row=to_row,
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        # تجواز الأميات
        if school == "أمية":
            continue

        if status == "خاتمة":
            khatemah(dar, student, id, track, level)
            continue
        else:
            test_template = load_workbook(
                filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                if str(track) == "1-ب" and i == 1:
                    continue
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['G4'] = track
            sheet['G5'] = level
            sheet['I4'] = id

        # file path creation
        isExist = os.path.exists(f"حفظ\{dar}\تعاهد")

        if not isExist:
            os.makedirs(f"حفظ\{dar}\تعاهد")

        # save
        test_template.save(filename=f"حفظ\{dar}\تعاهد\{student}.xlsx")
        print(f"{student} {id} is done - تعاهد")


# خاتمة

def khatemah(dar, student, id, track, level):
    test_template = load_workbook(
        filename="قوالب\قالب - استمارة تعاهد للدورات .xlsx")

    # file path creation
    sheet = test_template.active
    sheet['C4'] = student
    sheet['C5'] = dar
    sheet['G4'] = track
    sheet['G5'] = level
    sheet['I4'] = id

    # file path creation
    isExist = os.path.exists(f"خاتمات\{dar}")

    if not isExist:
        os.makedirs(f"خاتمات\{dar}")

    # save
    test_template.save(filename=f"خاتمات\{dar}\{student}.xlsx")
    print(f"{student} {id} is done - خاتمة")


# الأميات حفظ


def hefeth_M(database, from_row, to_row):
    hefeth_sheet = database['حفظ']

    for value in hefeth_sheet.iter_rows(min_row=from_row,
                                        max_row=to_row,
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        # تجواز غير الأميات
        if school != "أمية":
            continue

        test_template = load_workbook(
            filename="قوالب\قالب - استمارة اختبار الأمهات والأميات.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['F5'] = level
            sheet['H4'] = id

        # file path creation
        isExist = os.path.exists(f"أميات\{dar}\{teacher}")

        if not isExist:
            os.makedirs(f"أميات\{dar}\{teacher}")

        # save
        test_template.save(filename=f"أميات\{dar}\{teacher}\{student}.xlsx")
        print(f"{student} {id} is done - أميات حفظ")


# الأميات تعاهد


def hefeth_M(database, from_row, to_row):
    hefeth_sheet = database['تعاهد']

    for value in hefeth_sheet.iter_rows(min_row=from_row,
                                        max_row=to_row,
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        # تجواز غير الأميات
        if school != "أمية":
            continue

        test_template = load_workbook(
            filename="قوالب\قالب - استمارة اختبار الأمهات والأميات.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['F5'] = level
            sheet['H4'] = id

        # file path creation
        isExist = os.path.exists(f"أميات\{dar}\{teacher}\تعاهد")

        if not isExist:
            os.makedirs(f"أميات\{dar}\{teacher}\تعاهد")

        # save
        test_template.save(
            filename=f"أميات\{dar}\{teacher}\تعاهد\{student}.xlsx")
        print(f"{student} {id} is done - أميات تعاهد")


# ورقة التلقين


def talqeen(database, from_row, to_row):
    talqeen_sheet = database['منهج التلقين']

    for value in talqeen_sheet.iter_rows(min_row=from_row,
                                         max_row=to_row,
                                         min_col=1,
                                         max_col=8,
                                         values_only=True):
        (dar, halagah, teacher, student, id, school, track, level) = value

        test_template = load_workbook(
            filename="قوالب\قالب - استمارة اختبار التلقين التعليمي.xlsx")

        # fill cells
        sheet = test_template.active
        sheet['B4'] = student
        sheet['B5'] = dar
        sheet['D4'] = track
        sheet['D5'] = level
        sheet['F4'] = id

        # file path creation
        isExist = os.path.exists(f"استمارات التلقين\{dar}")

        if not isExist:
            os.makedirs(f"استمارات التلقين\{dar}")

        # save
        test_template.save(filename=f"استمارات التلقين\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - تلقين")


# ورقة التلاوة


def telawah(database, from_row, to_row):
    telawah_sheet = database['منهج التلاوة']

    for value in telawah_sheet.iter_rows(min_row=from_row,
                                         max_row=to_row,
                                         min_col=1,
                                         max_col=8,
                                         values_only=True):
        (dar, halagah, teacher, student, id, school, track, level) = value

        test_template = load_workbook(
            filename="قوالب\استمارات اختبارات منهج التلاوة.xlsx")

        # delete unwanted sheets
        for i in range(1, 9):
            if test_template[str(i)].title != str(level):
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template[str(level)]
        print(sheet)
        sheet['C4'] = student
        sheet['C5'] = dar
        sheet['G4'] = id

        # file path ceartion
        isExist = os.path.exists(f"استمارات التلاوة\{dar}")

        if not isExist:
            os.makedirs(f"استمارات التلاوة\{dar}")

        # save
        test_template.save(filename=f"استمارات التلاوة\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - تلاوة")


print('Done :D')
