from openpyxl import *
import os
import itertools

tests = os.listdir("test_templates")
database = load_workbook("database\قاعدة الفصل الأول - ١٤٤٤هـ.xlsx")


for test in tests:
    student_test = load_workbook("test_templates/" + test, data_only=True)
    sheet = student_test.active
    student_name = sheet["C4"].value
    student_track = sheet.title
    student_id = sheet["G4"].value if student_track == '6' else sheet["I4"].value

    print("Loking for:", student_name, student_id)

    hefeth_sheet = database["حفظ"]

    found = False
    for row in hefeth_sheet.iter_rows(min_row=2, max_row=len(list(hefeth_sheet.rows))):
        if row[3].value == student_name and row[4].value == student_id:
            print("we found her :D")
            found = True

            print("now we'r copying her marks")

            if student_track == '1':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=20, column=9).value
                print(hefeth_sheet.cell(row=row[1].row, column=10).value)
                for (i, j) in itertools.zip_longest(range(12, 45), range(27, 76, 6)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            elif student_track == '2':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=15, column=9).value
                for (i, j) in itertools.zip_longest(range(11, 44), range(21, 112, 5)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            elif student_track == '3':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=14, column=9).value
                for (i, j) in itertools.zip_longest(range(11, 44), range(19, 132, 4)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            elif student_track == '4':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=14, column=9).value
                for (i, j) in itertools.zip_longest(range(11, 44), range(19, 216, 4)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            elif student_track == '5':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=16, column=9).value
                for (i, j) in itertools.zip_longest(range(11, 44), range(23, 125, 6)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            elif student_track == '6':
                hefeth_sheet.cell(row=row[1].row, column=11).value = sheet.cell(
                    row=13, column=9).value
                for (i, j) in itertools.zip_longest(range(11, 44), range(18, 135, 4)):
                    if (not isinstance(sheet.cell(row=j, column=9).value, int)) and (not isinstance(sheet.cell(row=j, column=9).value, float)):
                        break
                    hefeth_sheet.cell(row=row[1].row, column=i).value = sheet.cell(
                        row=j, column=9).value

            print("copying is done :)")
            database.save(filename="database\قاعدة الفصل الأول - ١٤٤٤هـ.xlsx")

            # move to done file
            isExist = os.path.exists("done")
            if not isExist:
                os.makedirs("done")

            os.rename(f"test_templates\{test}", f"done\{test}")

    if not found:
        isExist = os.path.exists("not_found")
        if not isExist:
            os.makedirs("not_found")

        print("Sorry we couldn't find her :(")
        os.rename(f"test_templates\{test}", f"not_found\{test}")
