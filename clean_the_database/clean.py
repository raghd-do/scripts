from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# فتح القاعدة
database = load_workbook(filename="القاعدة\قاعدة الفصل الأول - ١٤٤٤هـ.xlsx")


def delete_unwanted_spaces(text):

    if text[0] == " ":
        text = text[1:]
    if text[-1] == " ":
        text = text[:-1]

    text = text.replace("   ", " ")
    text = text.replace("  ", " ")

    # تصحيح الأسماء التعبيدية
    text = text.replace("عبدالله", "عبد الله")
    text = text.replace("عبدالعزيز", "عبد العزيز")
    text = text.replace("عبدالرحمن", "عبد الرحمن")
    text = text.replace("عبدالخالق", "عبد الخالق")
    text = text.replace("عبدالقادر", "عبد القادر")
    text = text.replace("عبدالحق", "عبد الحق")
    text = text.replace("عبدالواحد", "عبد الواحد")
    text = text.replace("عبدالكريم", "عبد الكريم")
    text = text.replace("عبدالغني", "عبد الغني")
    text = text.replace("عبدالاله", "عبد الإله")

    # تعديل التاء المربوطة
    text = text.replace("منيره", "منيرة")
    text = text.replace("فاطمه", "فاطمة")
    text = text.replace("حصه", "حصة")
    text = text.replace("ساره", "سارة")
    text = text.replace("صفيه", "صفية")
    text = text.replace("رقيه", "رقية")
    text = text.replace("هيله", "هيلة")
    text = text.replace("ميمونه", "ميمونة")
    text = text.replace("لولوه", "لولوة")
    text = text.replace("مزنه", "مزنة")
    text = text.replace("بدريه", "بدرية")
    text = text.replace("شريفه", "شريفة")
    text = text.replace("اميه", "أمية")
    text = text.replace("أميه", "أمية")

    # همزة القطع والوصل
    text = text.replace("احمد", "أحمد")
    text = text.replace("ابراهيم", "إبراهيم")
    text = text.replace("امل", "أمل")
    text = text.replace("الزأمل", "الزامل")
    text = text.replace("اريج", "أريج")

    return text


def find_exact(database, qaid):
    hefeth_sheet = database['حفظ']
    taahod_sheet = database['تعاهد']
    monqatiaah = database['منقطعات']
    moatatherat = database['المعتذرات']

    delete = PatternFill(fgColor="ff0000", fill_type="solid")

    for row in hefeth_sheet.iter_rows(min_row=141,
                                      max_row=len(list(hefeth_sheet.rows)),
                                      min_col=1,
                                      max_col=9,
                                      ):
        if (qaid[1].value == row[1].value and qaid[3].value == row[3].value and qaid[4].value == row[4].value):
            qaid[3].fill == delete

    for row in taahod_sheet.iter_rows(min_row=2,
                                      max_row=len(list(taahod_sheet.rows)),
                                      min_col=1,
                                      max_col=9,
                                      ):
        if (qaid[1].value == row[1].value and qaid[3].value == row[3].value and qaid[4].value == row[4].value):
            qaid[3].fill == delete

    for row in monqatiaah.iter_rows(min_row=2,
                                    max_row=len(list(monqatiaah.rows)),
                                    min_col=1,
                                    max_col=9,
                                    ):
        if (qaid[1].value == row[1].value and qaid[3].value == row[3].value and qaid[4].value == row[4].value):
            qaid[3].fill == delete

    for row in moatatherat.iter_rows(min_row=2,
                                     max_row=len(list(moatatherat.rows)),
                                     min_col=1,
                                     max_col=9,
                                     ):
        if (qaid[1].value == row[1].value and qaid[3].value == row[3].value and qaid[4].value == row[4].value):
            qaid[3].fill == delete

    database.save("cleaned.xlsx")

# ورقة الحفظ


def clean_hefeth(database):
    hefeth_sheet = database['حفظ']
    taahod_sheet = database['تعاهد']
    monqatiaah = database['منقطعات']
    monqatiaah_qaid = database['منقطعات القيد']
    moatatherat = database['المعتذرات']

    from_qaid = PatternFill(fgColor="ffe699", fill_type="solid")
    new_student = PatternFill(fgColor="5b9bd5", fill_type="solid")
    not_found = PatternFill(fgColor="ffc000", fill_type="solid")
    nead_attention = PatternFill(fgColor="ffff00", fill_type="solid")
    delete = PatternFill(fgColor="ff0000", fill_type="solid")
    moatatherat_color = PatternFill(fgColor="bf8f00", fill_type="solid")

    moatatherat_ids = [c.value for c in moatatherat['E']]
    moatatherat_names = [c.value for c in moatatherat['D']]

    monqatiaah_qaid_ids = [c.value for c in monqatiaah_qaid['E']]
    monqatiaah_qaid_names = [c.value for c in monqatiaah_qaid['D']]

    # clean students' names hefeth
    for row in hefeth_sheet.iter_rows(min_row=2,
                                      max_row=len(list(hefeth_sheet.rows)),
                                      min_col=1,
                                      max_col=9,
                                      ):
        row[3].value = delete_unwanted_spaces(row[3].value)

        if row[3].value in moatatherat_names and row[4].value in moatatherat_ids:
            row[3].fill = moatatherat_color

        if row[3].value in monqatiaah_qaid and row[4].value in monqatiaah_qaid:
            row[3].fill = delete

    # compare and modify
    for qaid in hefeth_sheet.iter_rows(min_row=2,
                                       max_row=86,
                                       min_col=1,
                                       max_col=9,
                                       ):
        (qaid_dar, qaid_halagah, qaid_teacher, qaid_student, qaid_id,
         qaid_school, qaid_track, qaid_level, qaid_status) = qaid

        for target in hefeth_sheet.iter_rows(min_row=87, max_row=len(list(hefeth_sheet.rows)), min_col=1, max_col=9):
            (target_dar, target_halagah, target_teacher, target_student, target_id,
             target_school, target_track, target_level, target_status) = target

            # delete exact data
            if qaid_student.value == target_student.value and qaid_id.value == target_id.value:
                # hefeth_sheet.delete_rows(qaid_halagah.row)
                qaid_student.fill = delete

        for target in taahod_sheet.iter_rows(min_row=2, max_row=243, min_col=1, max_col=9):
            (target_dar, target_halagah, target_teacher, target_student, target_id,
             target_school, target_track, target_level, target_status) = target

            # delete exact data
            if qaid_student.value == target_student.value and qaid_id.value == target_id.value:
                # hefeth_sheet.delete_rows(qaid_halagah.row)
                qaid_student.fill = delete

        #     # color unsimilar data
        #     if qaid_halagah.value != target_halagah.value and qaid_student.value == target_student.value and qaid_id.value == target_id.value:
        #         if target_halagah.value == None or target_halagah.value == "":
        #             target_halagah.value = qaid_halagah.value
        #             # hefeth_sheet.delete_rows(qaid_halagah.row)
        #         else:
        #             qaid_halagah.fill = nead_attention
        #             target_halagah.fill = nead_attention
        #             continue

        # if qaid_student.value != target_student.value or qaid_id.value != target_id.value:
        #     qaid_student.fill = nead_attention
        #     target_student.fill = nead_attention
        #     continue

        # qaid_student.fill = new_student

    database.save("cleaned.xlsx")


clean_hefeth(database)
