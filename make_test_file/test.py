from openpyxl import Workbook, load_workbook
import os
from clean_spaces import *

# فتح القاعدة
database = load_workbook(
    filename="القاعدة/قاعدة الفصل الأول - ١٤٤٤هـ.xlsx")


# ورقة الحفظ بنات


def hefeth(database):
    hefeth_sheet = database['حفظ']

    if len(list(hefeth_sheet)) < 2:
        print("less than 2 rows")
        return

    print("في الحفظ: #", len(list(hefeth_sheet.rows)) - 1)

    for value in hefeth_sheet.iter_rows(min_row=2,
                                        max_row=len(list(hefeth_sheet.rows)),
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        dar = delete_unwanted_spaces(dar)
        teacher = delete_unwanted_spaces(teacher)
        student = delete_unwanted_spaces(student)

        # تجواز الأميات
        if school == "أميات":
            hefeth_M(value)
            continue

        test_template = load_workbook(
            filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                if str(track) == "1-ب" and i == 1:
                    continue
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = delete_unwanted_spaces(student)
        sheet['C5'] = delete_unwanted_spaces(dar)

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['G4'] = track
            sheet['G5'] = level
            sheet['I4'] = id

        # file path creation
        isExist = os.path.exists(f"حفظ\{dar}")

        if not isExist:
            os.makedirs(f"حفظ\{dar}")

        # save
        test_template.save(filename=f"حفظ\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - حفظ")


# ورقة التعاهد بنات


def taahod(database):
    taahod_sheet = database['تعاهد']

    if len(list(taahod_sheet)) < 2:
        print("less than 2 rows")
        return

    print("#", len(list(taahod_sheet.rows)) - 1)

    for value in taahod_sheet.iter_rows(min_row=2,
                                        max_row=len(list(taahod_sheet.rows)),
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        dar = delete_unwanted_spaces(dar)
        teacher = delete_unwanted_spaces(teacher)
        student = delete_unwanted_spaces(student)

        # الأميات
        if school == "أمية":
            taahod_M(value)
            continue

        if status == "خاتمة تعاهد":
            khatemah(value)
            continue
        else:
            test_template = load_workbook(
                filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                if str(track) == "1-ب" and i == 1:
                    continue
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        print(student)
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['G4'] = track
            sheet['G5'] = level
            sheet['I4'] = id

        # file path creation
        isExist = os.path.exists(f"حفظ\{dar}\تعاهد")

        if not isExist:
            os.makedirs(f"حفظ\{dar}\تعاهد")

        # save
        test_template.save(filename=f"حفظ\{dar}\تعاهد\{student}.xlsx")
        print(f"{student} {id} is done - تعاهد")

# تعاهد الدورات


def taahod_dowrat(database):
    taahod_sheet = database['تعاهد']

    if len(list(taahod_sheet)) < 2:
        print("less than 2 rows")
        return

    print("في تعاهد الدورات: #", len(list(taahod_sheet.rows)) - 1)

    for value in taahod_sheet.iter_rows(min_row=2,
                                        max_row=len(list(taahod_sheet.rows)),
                                        min_col=1,
                                        max_col=9,
                                        values_only=True):
        (dar, halagah, teacher, student, id, school, track, level, status) = value

        # Open test template
        test_template = load_workbook(
            filename="قوالب/قالب - استمارة تعاهد للدورات .xlsx")

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['G4'] = track
            sheet['G5'] = level
            sheet['I4'] = id
            sheet['I5'] = teacher

        # file path creation
        isExist = os.path.exists(f"{dar}")

        if not isExist:
            os.makedirs(f"{dar}")

        # save
        test_template.save(filename=f"{dar}/{student}.xlsx")
        print(f"{student} {id} is done - تعاهد")

# خاتمة


def khatemah(dar, student, id, track, level):
    test_template = load_workbook(
        filename="قوالب\قالب - استمارة تعاهد للدورات .xlsx")

    # file path creation
    sheet = test_template.active
    sheet['C4'] = student
    sheet['C5'] = dar
    sheet['G4'] = track
    sheet['G5'] = level
    sheet['I4'] = id

    # file path creation
    isExist = os.path.exists(f"خاتمات\{dar}")

    if not isExist:
        os.makedirs(f"خاتمات\{dar}")

    # save
    test_template.save(filename=f"خاتمات\{dar}\{student}.xlsx")
    print(f"{student} {id} is done - خاتمة")


# الأميات حفظ


def hefeth_M(value):
    (dar, halagah, teacher, student, id, school, track, level, status) = value

    dar = delete_unwanted_spaces(dar)
    teacher = delete_unwanted_spaces(teacher)
    student = delete_unwanted_spaces(student)

    test_template = load_workbook(
        filename="قوالب\قالب - استمارة اختبار الأمهات والأميات.xlsx")

    # delete unwanted sheets
    for i in range(1, 7):
        if test_template[str(i)].title != str(track):
            if str(track) == "1-ب" and i == 1:
                    continue
            test_template.remove(test_template[str(i)])

    # fill cells
    sheet = test_template.active
    sheet['C4'] = student
    sheet['C5'] = dar

    if track == 6:
        sheet['E5'] = level
        sheet['G4'] = id
    else:
        sheet['F5'] = level
        sheet['H4'] = id

    # file path creation
    isExist = os.path.exists(f"أميات\{dar}\{teacher}")

    if not isExist:
        os.makedirs(f"أميات\{dar}\{teacher}")

    # save
    test_template.save(filename=f"أميات\{dar}\{teacher}\{student}.xlsx")
    print(f"{student} {id} is done - أميات حفظ")


# الأميات تعاهد


def taahod_M(value):
    (dar, halagah, teacher, student, id, school, track, level, status) = value

    dar = delete_unwanted_spaces(dar)
    teacher = delete_unwanted_spaces(teacher)
    student = delete_unwanted_spaces(student)

    if status == "خاتمة تعاهد":
        khatemah(dar, student, id, track, level)
    else:
        test_template = load_workbook(
            filename="قوالب\قالب - استمارة اختبار الأمهات والأميات.xlsx")

        # delete unwanted sheets
        for i in range(1, 7):
            if test_template[str(i)].title != str(track):
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = dar

        if track == 6:
            sheet['E5'] = level
            sheet['G4'] = id
        else:
            sheet['F5'] = level
            sheet['H4'] = id

        # file path creation
        isExist = os.path.exists(f"أميات\{dar}\{teacher}\تعاهد")

        if not isExist:
            os.makedirs(f"أميات\{dar}\{teacher}\تعاهد")

        # save
        test_template.save(
            filename=f"أميات\{dar}\{teacher}\تعاهد\{student}.xlsx")
        print(f"{student} {id} is done - أميات تعاهد")


# ورقة التلقين


def talqeen(database):
    talqeen_sheet = database['منهج التلقين']

    for value in talqeen_sheet.iter_rows(min_row=3,
                                         max_row=len(list(talqeen_sheet.rows)),
                                         min_col=1,
                                         max_col=8,
                                         values_only=True):
        (dar, halagah, teacher, student, id, school, track, level) = value

        test_template = load_workbook(
            filename="قوالب\قالب - استمارة اختبار التلقين التعليمي.xlsx")

        # fill cells
        sheet = test_template.active
        sheet['B4'] = student
        sheet['B5'] = dar
        sheet['D4'] = track
        sheet['D5'] = level
        sheet['F4'] = id

        # file path creation
        isExist = os.path.exists(f"استمارات التلقين\{dar}")

        if not isExist:
            os.makedirs(f"استمارات التلقين\{dar}")

        # save
        test_template.save(filename=f"استمارات التلقين\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - تلقين")


# ورقة التلاوة


def telawah(database):
    telawah_sheet = database['منهج التلاوة']

    for value in telawah_sheet.iter_rows(min_row=3,
                                         max_row=len(list(telawah_sheet.rows)),
                                         #  max_row=12,
                                         min_col=1,
                                         max_col=8,
                                         values_only=True):
        (dar, halagah, teacher, student, id, school, track, level) = value

        test_template = load_workbook(
            filename="قوالب\استمارات اختبارات منهج التلاوة.xlsx")

        # delete unwanted sheets
        for i in range(1, 9):
            if test_template[str(i)].title != str(level):
                test_template.remove(test_template[str(i)])

        # fill cells
        sheet = test_template[str(level)]
        sheet['C4'] = student
        sheet['C5'] = dar
        sheet['G4'] = id
        sheet['G5'] = teacher

        # file path ceartion
        isExist = os.path.exists(f"استمارات التلاوة\{dar}")

        if not isExist:
            os.makedirs(f"استمارات التلاوة\{dar}")

        # save
        test_template.save(filename=f"استمارات التلاوة\{dar}\{student}.xlsx")
        print(f"{student} {id} is done - تلاوة")


# ورقة خاتمات

def katemah(database):
    sheet = database['خاتمات']

    if len(list(sheet)) < 2:
        print("less than 2 rows")
        return

    print("في الخاتمات: #", len(list(sheet.rows)) - 1)

    for value in sheet.iter_rows(min_row=2,
                                        max_row=len(list(sheet.rows)),
                                        min_col=1,
                                        max_col=2,
                                        values_only=True):
        (student, id) = value

        test_template = load_workbook(
            filename="./قوالب/استمارة كامل القرآن (اختبار الختمة).xlsx")

        # Get the names of the first three sheets
        sheet_names = test_template.sheetnames[:3]

        # fill cells in the first three sheets
        for sheet_name in sheet_names:
            sheet = test_template[sheet_name]
            sheet["C2"] = delete_unwanted_spaces(student)

        # file path creation
        isExist = os.path.exists(f"خاتمات")

        if not isExist:
            os.makedirs(f"خاتمات")

        # save
        test_template.save(filename=f"خاتمات\{student}.xlsx")
        print(f"{student} {id} is done - خاتمات")

# make test template for one


def make_test_template_hefeth(value):
    (dar, halagah, teacher, student, id, school, track, level, status) = value

    # تجواز الأميات
    if school == "أمية":
        hefeth_M(dar, teacher, student, id, track, level)
        return

    test_template = load_workbook(
        filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

    # delete unwanted sheets
    for i in range(1, 7):
        if test_template[str(i)].title != str(track):
            if str(track) == "1-ب" and i == 1:
                continue
            test_template.remove(test_template[str(i)])

    # fill cells
    sheet = test_template.active
    sheet['C4'] = student
    sheet['C5'] = dar

    if track == 6:
        sheet['E5'] = level
        sheet['G4'] = id
    else:
        sheet['G4'] = track
        sheet['G5'] = level
        sheet['I4'] = id

    # file path creation
    isExist = os.path.exists(f"حفظ\{dar}")

    if not isExist:
        os.makedirs(f"حفظ\{dar}")

    # save
    test_template.save(filename=f"حفظ\{dar}\{student}.xlsx")
    print(f"{student} {id} is done - حفظ")


def make_test_template_taahod(value):
    (dar, halagah, teacher, student, id, school, track, level, status) = value

    # الأميات
    if school == "أمية":
        taahod_M(value)
        return

    if status == "خاتمة تعاهد":
        khatemah(value)
        return
    else:
        test_template = load_workbook(
            filename="قوالب\قالب - استمارات اختبارات المناهج.xlsx")

    # delete unwanted sheets
    for i in range(1, 7):
        if test_template[str(i)].title != str(track):
            if str(track) == "1-ب" and i == 1:
                continue
            test_template.remove(test_template[str(i)])

    # fill cells
    sheet = test_template.active
    sheet['C4'] = student
    sheet['C5'] = dar

    if track == 6:
        sheet['E5'] = level
        sheet['G4'] = id
    else:
        sheet['G4'] = track
        sheet['G5'] = level
        sheet['I4'] = id

    # file path creation
    isExist = os.path.exists(f"حفظ\{dar}\تعاهد")

    if not isExist:
        os.makedirs(f"حفظ\{dar}\تعاهد")

    # save
    test_template.save(filename=f"حفظ\{dar}\تعاهد\{student}.xlsx")
    print(f"{student} {id} is done - تعاهد")

# استمارات اختبار قبول المعهد - اختبار 5 أجزاء


# def Maahad_Test(database):
    Maahad_sheet = database['المعهد']

    for value in Maahad_sheet.iter_rows(min_row=2,
                                        max_row=len(list(Maahad_sheet.rows)),
                                        min_col=1,
                                        max_col=3,
                                        values_only=True):
        (student, track, chapter) = value

        test_template = load_workbook(
            filename="قوالب\قالب - استمارات قبول المعهد اختبار 5 أجزاء.xlsx")

        # fill cells
        sheet = test_template.active
        sheet['C4'] = student
        sheet['C5'] = track
        sheet['D6'] = chapter

        # file path ceartion
        isExist = os.path.exists(f"استمارات المعهد\{track}")

        if not isExist:
            os.makedirs(f"استمارات المعهد\{track}")

        # save
        test_template.save(filename=f"استمارات المعهد\{track}\{student}.xlsx")
        print(f"{student} is done - المعهد")

# حفظ
# hefeth(database)

# تعاهد
# taahod(database)


# تعاهد دورات
# taahod_dowrat(database)


# تلاوة
# telawah(database)

# تلقين
# talqeen(database)

# المعهد
# Maahad_Test(database)

# خاتمات
katemah(database)

print('Done :D')
